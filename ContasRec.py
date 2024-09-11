import pandas as pd
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox
import random
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def calcular_novas_datas(data_inicio, data_final):
    datas = []
    data_inicio += relativedelta(months=1)
    while data_inicio <= data_final:
        datas.append(data_inicio)
        data_inicio += relativedelta(months=1)
    return datas
def gerar_numero_referencia():
    return random.randint(100000000, 999999999)
def processar_arquivo(file_path, output_file_path):
    df = pd.read_excel(file_path)
    df['Data de lançamento'] = pd.to_datetime(df['Data de lançamento'], format='%d/%m/%Y')
    df['Vencimento líquido'] = pd.to_datetime(df['Vencimento líquido'], format='%d/%m/%Y')
    data_atual = datetime.now()
    novos_registros = []
    for _, row in df.iterrows():
        if data_atual + (timedelta(days=30)) >= row['Vencimento líquido']:
            data_atual += timedelta(days=30)
            novos_registros.append(row)
            novas_datas = calcular_novas_datas(data_atual,row['Vencimento líquido'] + relativedelta(months=4))
            for nova_data in novas_datas:
                novo_registro = row.copy()
                novo_registro['Vencimento líquido'] = nova_data
                novo_registro['Referência'] = gerar_numero_referencia()  
                novos_registros.append(novo_registro)
        elif data_atual + (timedelta(days=60)) >= row['Vencimento líquido'] - (timedelta(days=30)):
            novos_registros.append(row)  
            novas_datas = calcular_novas_datas(data_atual, row['Vencimento líquido']) 
               
            for nova_data in novas_datas: 
                novo_registro = row.copy() 
                novo_registro['Vencimento líquido'] = nova_data  
                novo_registro['Referência'] = gerar_numero_referencia()  
                novos_registros.append(novo_registro)    
        else:    
            novos_registros.append(row)  
            novas_datas = calcular_novas_datas(data_atual, row['Vencimento líquido']-timedelta(days=30))  
               
            for nova_data in novas_datas: 
                novo_registro = row.copy()  
                novo_registro['Vencimento líquido'] = nova_data  
                novo_registro['Referência'] = gerar_numero_referencia()  
                novos_registros.append(novo_registro) 
            
    df_novo = pd.DataFrame(novos_registros)
    df_novo.to_excel(output_file_path, index=False)

    workbook = load_workbook(output_file_path)
    sheet = workbook.active
  
    for column in sheet.columns:
       max_length = 0
       column_letter = get_column_letter(column[0].column)
       for cell in column:
           try:
               if len(str(cell.value)) > max_length:
                   max_length = len(str(cell.value))
           except:
               pass
       adjusted_width = (max_length + 2)
       sheet.column_dimensions[column_letter].width = adjusted_width
   
    for row in sheet.iter_rows():
       for cell in row:
           if isinstance(cell.value, datetime):
               cell.number_format = 'DD/mm/YYYY'
  
    workbook.save(output_file_path)
    messagebox.showinfo("Sucesso", f"Arquivo salvo com sucesso como {output_file_path}")
def selecionar_arquivo():
   file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
   if file_path:
       entrada_var.set(file_path)
def selecionar_pasta():
   folder_path = filedialog.askdirectory()
   if folder_path:
       saida_var.set(folder_path)
def executar():
   file_path = entrada_var.get()
   folder_path = saida_var.get()
   if not file_path or not folder_path:
       messagebox.showerror("Erro", "Por favor, selecione o arquivo de entrada e a pasta de saída.")
       return
   output_file_path = f"{folder_path}/Base a Receber Atualizada.xlsx"
   processar_arquivo(file_path, output_file_path)

root = tk.Tk()
root.title("Processador de Arquivos Excel")
entrada_var = tk.StringVar()
saida_var = tk.StringVar()
tk.Label(root, text="Selecione o arquivo de entrada:").grid(row=0, column=0, padx=10, pady=10)
tk.Entry(root, textvariable=entrada_var, width=50).grid(row=0, column=1, padx=10, pady=10)
tk.Button(root, text="Procurar", command=selecionar_arquivo).grid(row=0, column=2, padx=10, pady=10)
tk.Label(root, text="Selecione a pasta de saída:").grid(row=1, column=0, padx=10, pady=10)
tk.Entry(root, textvariable=saida_var, width=50).grid(row=1, column=1, padx=10, pady=10)
tk.Button(root, text="Procurar", command=selecionar_pasta).grid(row=1, column=2, padx=10, pady=10)
tk.Button(root, text="Executar", command=executar).grid(row=2, column=0, columnspan=3, pady=20)
root.mainloop()